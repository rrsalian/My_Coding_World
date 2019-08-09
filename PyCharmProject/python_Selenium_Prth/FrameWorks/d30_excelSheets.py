import openpyxl

book = openpyxl.load_workbook(filename="D:\python_Selenium\d30_excel.xlsx")
sheet = book.active

row = sheet.max_row
column = sheet.max_column

print("The max no. of rows:",row)
print("The max no. of columns:",column)

for r in range(1, row+1):
    for c in range(1, column+1):
        print(sheet.cell(row=r,column=c).value,end="\t\t")
    print()

#writing:
sheet.cell(row=7,column=1).value="pratik"
sheet.cell(row=7,column=2).value=16
sheet.cell(row=7,column=3).value=210000

book.save(filename="D:\python_Selenium\d30_excel.xlsx")

book.close()