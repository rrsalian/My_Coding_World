import openpyxl

book=openpyxl.load_workbook(filename="E:\SeleniumComponents\Cricket.xlsx")

sheet=book.active

row=sheet.max_row
Column=sheet.max_column

print("Row count :",row)
print("Column count :",Column)

print()

for r in range(1,row+1):
    for c in range(1,Column+1):
        print((str(sheet.cell(row=r,column=c).value)).center(15," "),end="\t\t")
    print()



book.close()