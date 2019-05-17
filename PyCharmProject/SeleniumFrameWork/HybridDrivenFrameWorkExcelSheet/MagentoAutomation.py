import time

import openpyxl
from selenium import webdriver

book=openpyxl.load_workbook(filename="E:\SeleniumComponents\MagentoLogin.xlsx")

sheet=book.active

row=sheet.max_row
Column=sheet.max_column

print("Row count :",row)
print("Column count :",Column)

for r in range(2,row+1):
    action=sheet.cell(row=r,column=3).value
    if(action=="open"):
        chrome=webdriver.Chrome(executable_path="E:\SeleniumComponents\chromedriver.exe")
        chrome.maximize_window()
        chrome.implicitly_wait(20)
    elif(action=="navigator"):
        chrome.get(url=sheet.cell(row=r,column=5).value)
    elif(action=="click"):
        chrome.find_element_by_xpath(xpath=sheet.cell(row=r,column=4).value).click()
    elif(action=="type"):
        chrome.find_element_by_xpath(xpath=sheet.cell(row=r,column=4).value).send_keys(sheet.cell(row=r,column=5).value)
    elif(action=="quit"):
        time.sleep(7)
        chrome.quit()
    else:
        print("Invalid action")
