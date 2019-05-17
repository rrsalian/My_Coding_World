import openpyxl

book=openpyxl.load_workbook(filename="E:\SeleniumComponents\Cricket.xlsx")

sheet=book.active

row=sheet.max_row
Column=sheet.max_column

print("Row count :",row)
print("Column count :",Column)

print()

sheet.cell(row=7,column=1).value="Raina"
sheet.cell(row=7,column=2).value=8
sheet.cell(row=7,column=3).value=40000


book.save(filename="E:\SeleniumComponents\Cricket.xlsx")




book.close()